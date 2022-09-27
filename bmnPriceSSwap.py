import json
from websocket import create_connection
import openpyxl
from time import sleep
import datetime

from datetime import datetime as dt
ws = create_connection("wss://api.sideswap.io/json-rpc-ws")
ws.send('{"id":1, "method": "market_data_subscribe", "params": {"asset": "11f91cb5edd5d0822997ad81f068ed35002daec33986da173461a8427ac857e1"}}')

# while True:
result =  ws.recv()
# print ("Received '%s'" % result)
with open('BMN1.json', 'w') as f:
    f.write(result)
f.close()
ws.close()

sleep(2)
print("esperando a que el fichero se cierre correctamente")

def extractBMN():
    f = open('BMN1.json',)
    datos = json.load(f)
    closes = []
    times = []
    for i in datos['result']['data']:
        d = datetime.datetime.strptime(i['time'], "%Y-%m-%d")
        d = d.strftime("%d-%m-%Y")
        print(d, i['close'])
        times.append(d)
        closes.append(i['close'])
    f.close()
    return closes, times

def writeBMN():
    cierres, fechas = extractBMN()
    wb = openpyxl.load_workbook('pruebas.xlsx')
    sheet = wb['Sheet2']
    for i in range(len(cierres)):
        sheet.cell(row=i+2, column=5).value = fechas[i]
        sheet.cell(row=i+2, column=6).value = cierres[i]
    wb.save('pruebas.xlsx')

writeBMN()