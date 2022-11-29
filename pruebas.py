import openpyxl
import requests
import pandas as pd
import json
# from datos import *
from datetime import datetime, timedelta
# from datetime import datetime as dt

d = datetime.today() - timedelta(days=0)
now = str(d.strftime('%d-%m-%Y'))

def dataframeExtract(que):
    url = 'https://api.blockchain.info/charts/'
    url = url + que
    res = requests.get(url, params={'timespan': '10weeks'})
    df = pd.read_json(res.text, convert_dates=['t'])
    return df


def extractHR():
    fes = []
    fes2 = []
    fes3 = []
    hrs = []
    rev = []
    pre = []
    df = dataframeExtract('hash-rate')
    dfI = dataframeExtract('miners-revenue')
    dfP = dataframeExtract('market-price')

    res1 = df['values']
    res2 = dfI['values']
    res3 = dfP['values']
    for i in range(len(res1)):
        res1[i]['x'] = pd.to_datetime(res1[i]['x'], unit='s')
        d = res1[i]['x'].strftime("%d-%m-%Y")
        fes.append(d)
        hrs.append(df['values'][i]['y'])
    for i in range(len(res2)):
        res2[i]['x'] = pd.to_datetime(res2[i]['x'], unit='s')
        d = res2[i]['x'].strftime("%d-%m-%Y")
        fes2.append(d)
        rev.append(dfI['values'][i]['y'])
    for i in range(len(res3) - 1):
        res3[i]['x'] = pd.to_datetime(res3[i]['x'], unit='s')
        d = res3[i]['x'].strftime("%d-%m-%Y")
        fes3.append(d)
        pre.append(dfP['values'][i]['y'])
    return fes, hrs, rev, pre

# hr = extractHR()
#
# def extractIngresos():
#     fes = []
#     rev = []
#     df = dataframeExtract('miners-revenue')
#     res = df['values']
#     for i in range(len(res)):
#         res[i]['x'] = pd.to_datetime(res[i]['x'], unit='s')
#         d = res[i]['x'].strftime("%d-%m-%Y")
#         fes.append(d)
#         rev.append(df['values'][i]['y'])
#     return fes, rev
#
# ingresos = extractIngresos()
#
# def extractPrecios():
#     fes = []
#     pre = []
#     df = dataframeExtract('market-price')
#     res = df['values']
#     for i in range(len(res)):
#         res[i]['x'] = pd.to_datetime(res[i]['x'], unit='s')
#         d = res[i]['x'].strftime("%d-%m-%Y")
#         fes.append(d)
#         pre.append(df['values'][i]['y'])
#     return fes, pre

# precios = extractPrecios()

print(now)
# now = pd.to_datetime(now, format = "%d/%m/%Y", infer_datetime_format=True)
def writeHR():
    fechas, hashrates = extractHR()
    wb = openpyxl.load_workbook('pruebas.xlsx')
    sheet = wb['Sheet1']
    for i in range(len(fechas)):
        sheet.cell(row=i+2, column=5).value = fechas[i]
        sheet.cell(row=i+2, column=6).value = hashrates[i]
    wb.save('pruebas.xlsx')

def writeRev():
    fechas, revenues = extractIngresos()
    wb = openpyxl.load_workbook('pruebas.xlsx')
    sheet = wb['Sheet1']
    for i in range(len(fechas)):
        sheet.cell(row=i+2, column=8).value = fechas[i]
        sheet.cell(row=i+2, column=9).value = revenues[i]
    wb.save('pruebas.xlsx')

def writePre():
    fechas, precios = extractPrecios()
    wb = openpyxl.load_workbook('pruebas.xlsx')
    sheet = wb['Sheet1']
    for i in range(len(fechas)):
        sheet.cell(row=i+2, column=10).value = fechas[i]
        sheet.cell(row=i+2, column=11).value = precios[i]
    wb.save('pruebas.xlsx')

# writeHR()
# writeRev()
# writePre()

